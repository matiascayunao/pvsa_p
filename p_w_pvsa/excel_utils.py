from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy
from .models import ObjetoLugar

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_TITLE = PatternFill("solid", fgColor="CFE2F3")
FILL_PISO = PatternFill("solid", fgColor="E7F1FF")

FONT_TITLE = Font(bold=True,  size=14)
FONT_PISO= Font(bold=True, size=12)
FONT_HDR= Font(bold=True, size=10)
FONT_CELL = Font(size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

FILL_ESTADO_BUENO = PatternFill("solid", fgColor="C6EFCE")
FILL_ESTADO_PENDIENTE = PatternFill("solid", fgColor="FFF2CC")
FILL_ESTADO_MALO = PatternFill("solid", fgColor="F9CBAD")

def _safe_sheet_name(name: str ) -> str:
    bad = [":","\\","/","?", "*","[","]"]
    for ch in bad:
        name = name.replace(ch, " ")
    name = " ".join(name.split()).strip()
    return name[:31] if len(name) > 31 else name

def _unique_sheet_name(wb: workbook, base_name: str) -> str:
    name= _safe_sheet_name(base_name)
    if name not in wb.sheetnames:
        return name
    i = 2
    while True:
        suffix = f"({i})"
        cut = 31 - len(suffix)
        candidate = _safe_sheet_name(name[:cut] + suffix)
        if candidate not in wb.sheetnames:
            return candidate
        i += 1

def _set_col_widths(ws):
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 14

def _style_row(ws,row,c1,c2,fill=None, font=None, align=None):
    for c in range(c1,c2 +1 ):
        cell = ws.cell(row=row, column=c)
        cell.border = BORDER
        if fill:
            cell.fill = copy(fill)
        if font:
            cell.font = copy(font)
        if align:
            cell.alignment = copy(align)

def _write_lugar_block(ws, start_row, lugar, objetos_qs):
    ws.merge_cells(start_row=start_row, start_column= 1, end_row=start_row, end_column=6)
    title_font= Font(bold=True, size=12)
    t = ws.cell(start_row, 1, lugar.nombre_del_lugar)
    t.font = title_font
    t.alignment = CENTER
    _style_row(ws, start_row,1,6,fill=FILL_TITLE, font=title_font, align=CENTER)

    headers = ["Objeto", "Tipo", "Cantidad", "Estado", "Detalle", "Fecha"]
    hr = start_row + 1
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(hr, i ,h)
        cell.font = FONT_HDR
        cell.alignment = CENTER
        cell.fill = FILL_TITLE
        cell.border = BORDER

    r = hr + 1 
    count = 0
    for ol in objetos_qs:
        count += 1
        objeto_nombre = ol.tipo_de_objeto.objeto.nombre_del_objeto
        ### tipo_txt = f"{ol.tipo_de_objeto.marca}  -  {ol.tipo_de_objeto.material}"
        marca = (ol.tipo_de_objeto.marca or "").strip()
        material = (ol.tipo_de_objeto.material or "").strip()
        partes = [p for p in (marca, material)if p] 
        tipo_txt= " - ".join(partes) if partes else "-"

        ws.cell (r,1,objeto_nombre).font = FONT_CELL
        ws.cell (r,2,tipo_txt).font = FONT_CELL
        ws.cell(r,3,ol.cantidad).font = FONT_CELL

        estado = ol.get_estado_display()
        estado_cell = ws.cell(r,4, estado)
        estado_cell.font = FONT_CELL
        
        if estado == "Bueno":
            estado_cell.fill = FILL_ESTADO_BUENO
        elif estado == "Pendiente":
            estado_cell.fill = FILL_ESTADO_PENDIENTE
        elif estado == "Malo":
            estado_cell.fill = FILL_ESTADO_MALO

        ws.cell (r,5,ol.detalle or "-").font = FONT_CELL
        cfecha = ws.cell(r,6,ol.fecha)
        cfecha.number_format = "dd/mm/yyyy"
        cfecha.font = FONT_CELL

        for c in range(1,7):
            ws.cell(r,c).border = BORDER
            ws.cell(r,c).alignment = LEFT if c in (1,2,5) else CENTER
        r += 1
    if count ==0:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        cell = ws.cell(r,1,"Sin objetos registrados")
        cell.font = Font(italic=True, size=10)
        cell.alignment = CENTER
        _style_row(ws,r,1,6,align=CENTER)
        r+=1

    return r + 1
def build_excel_sectores(ubicaciones_qs):
    wb = Workbook()
    wb.remove(wb.active)
    

    for ub in ubicaciones_qs:
        sheet_name= _unique_sheet_name(wb,f"{ub.sector.sector} - {ub.ubicacion}")
        ws = wb.create_sheet(title=sheet_name)
        _set_col_widths(ws)

        ws.merge_cells("A1:F1")
        ws["A1"] = f"Sector: {ub.sector.sector} | Ubicación: {ub.ubicacion}"

        ws["A1"].font = FONT_TITLE
        ws["A1"].alignment = CENTER
        _style_row(ws,1,1,6, fill=FILL_TITLE, font= FONT_TITLE, align=CENTER)
        row= 3

        pisos = ub.piso_set.all().order_by("piso")
        for p in pisos:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row,1,f"PISO {p.piso}")
            cell.font = FONT_PISO
            cell.alignment = LEFT
            _style_row(ws, row,1,6, fill = FILL_PISO, font=FONT_PISO, align=LEFT)
            row +=2

            lugares = p.lugar_set.all().order_by("id")
            for lugar in lugares:
                objetos = (ObjetoLugar.objects.filter(lugar=lugar).select_related("tipo_de_objeto__objeto").order_by("id"))
                row = _write_lugar_block(ws, row, lugar, objetos)
            row += 1
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()



    